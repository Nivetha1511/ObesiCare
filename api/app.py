from flask import Flask, render_template, request, jsonify, send_from_directory
from flask_cors import CORS
import tensorflow as tf
from keras.models import load_model
import pickle
import pandas as pd
import numpy as np
import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
import hashlib

# Resolve project paths from this file so app works from any launch directory.
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
FRONTEND_DIR = os.path.join(BASE_DIR, 'frontend')
CSS_DIR = os.path.join(FRONTEND_DIR, 'css')
JS_DIR = os.path.join(FRONTEND_DIR, 'js')
IMAGES_DIR = os.path.join(FRONTEND_DIR, 'images')

# Initialize Flask app
app = Flask(__name__, template_folder=FRONTEND_DIR, static_folder=FRONTEND_DIR, static_url_path='')
CORS(app)

# Load ML model and preprocessing objects
print("Loading model artifacts...")
MODEL_PATH = os.path.join(BASE_DIR, 'models', 'obesity_ann.keras')
SCALER_PATH = os.path.join(BASE_DIR, 'models', 'scaler.pkl')
ENCODER_PATH = os.path.join(BASE_DIR, 'models', 'target_encoder.pkl')
FEATURES_PATH = os.path.join(BASE_DIR, 'models', 'feature_columns.pkl')

CLASS_DISPLAY_MAP = {
    'Insufficient_Weight': 'Insufficient Weight',
    'Normal_Weight': 'Normal Weight',
    'Overweight_Level_I': 'Overweight Level I',
    'Overweight_Level_II': 'Overweight Level II',
    'Obesity_Type_I': 'Obesity Type I',
    'Obesity_Type_II': 'Obesity Type II',
    'Obesity_Type_III': 'Obesity Type III'
}

NUMERIC_CLASS_FALLBACK = {
    '0': 'Insufficient_Weight',
    '1': 'Normal_Weight',
    '2': 'Obesity_Type_I',
    '3': 'Obesity_Type_II',
    '4': 'Obesity_Type_III',
    '5': 'Overweight_Level_I',
    '6': 'Overweight_Level_II'
}

OBESITY_RISK_MAP = {
    'Insufficient_Weight': {'risk_level': 'Insufficient Weight', 'icon': '⬇️', 'color': '#3498db'},
    'Normal_Weight': {'risk_level': 'Normal Weight', 'icon': '✅', 'color': '#2ecc71'},
    'Overweight_Level_I': {'risk_level': 'Overweight Level I', 'icon': '⚠️', 'color': '#f39c12'},
    'Overweight_Level_II': {'risk_level': 'Overweight Level II', 'icon': '⚠️', 'color': '#e67e22'},
    'Obesity_Type_I': {'risk_level': 'Obesity Type I', 'icon': '🔴', 'color': '#e74c3c'},
    'Obesity_Type_II': {'risk_level': 'Obesity Type II', 'icon': '🔴', 'color': '#c0392b'},
    'Obesity_Type_III': {'risk_level': 'Obesity Type III', 'icon': '🔴', 'color': '#8b0000'}
}


def _pediatric_class_from_bmi(bmi):
    """Rule-based fallback for ages 10-15 where training data is sparse."""
    # Tuned to match observed BMI bands in age_10_15_raw.csv
    if bmi < 18.5:
        return 'Insufficient_Weight', 0
    if bmi < 21.0:
        return 'Normal_Weight', 1
    if bmi < 25.0:
        return 'Overweight_Level_I', 2
    if bmi < 27.0:
        return 'Overweight_Level_II', 3
    if bmi < 29.5:
        return 'Obesity_Type_I', 4
    if bmi < 31.5:
        return 'Obesity_Type_II', 5
    return 'Obesity_Type_III', 6


def _pediatric_probabilities(center_idx):
    """Generate smooth class probabilities centered around rule-based class."""
    base = np.array([2.0] * 7)
    base[center_idx] += 72.0
    if center_idx - 1 >= 0:
        base[center_idx - 1] += 12.0
    if center_idx + 1 <= 6:
        base[center_idx + 1] += 12.0
    if center_idx - 2 >= 0:
        base[center_idx - 2] += 4.0
    if center_idx + 2 <= 6:
        base[center_idx + 2] += 4.0
    probs = base / np.sum(base)
    return probs

try:
    model = load_model(MODEL_PATH)
    with open(SCALER_PATH, 'rb') as f:
        scaler = pickle.load(f)
    with open(ENCODER_PATH, 'rb') as f:
        target_encoder = pickle.load(f)
    with open(FEATURES_PATH, 'rb') as f:
        feature_columns = pickle.load(f)
    print("✓ Model artifacts loaded successfully")
except Exception as e:
    print(f"Error loading model: {e}")
    model = None

# User data file
USERS_FILE = os.path.join(BASE_DIR, 'user_data', 'users.xlsx')

def ensure_user_file():
    """Ensure user data file exists"""
    os.makedirs(os.path.dirname(USERS_FILE), exist_ok=True)
    if not os.path.exists(USERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'username'
        ws['B1'] = 'password_hash'
        ws['C1'] = 'created_at'
        wb.save(USERS_FILE)
        print("✓ Created user data file")

def hash_password(password):
    """Hash password using SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()

def user_exists(username):
    """Check if user exists in database"""
    try:
        wb = load_workbook(USERS_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[0].lower() == username.lower():
                return True
        return False
    except:
        return False

def verify_user(username, password):
    """Verify user credentials"""
    try:
        wb = load_workbook(USERS_FILE)
        ws = wb.active
        password_hash = hash_password(password)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[0].lower() == username.lower():
                if row[1] == password_hash:
                    return True
        return False
    except:
        return False

def add_user(username, password):
    """Add new user to database"""
    try:
        wb = load_workbook(USERS_FILE)
        ws = wb.active
        password_hash = hash_password(password)
        new_row = [username, password_hash, datetime.now().isoformat()]
        ws.append(new_row)
        wb.save(USERS_FILE)
        return True
    except Exception as e:
        print(f"Error adding user: {e}")
        return False

# Initialize user file on startup
ensure_user_file()

# ==================== Routes ====================

@app.route('/')
def index():
    """Home page"""
    return render_template('index.html')

@app.route('/login')
def login():
    """Login page"""
    return render_template('login.html')

@app.route('/signup')
def signup():
    """Sign up page"""
    return render_template('signup.html')

@app.route('/form')
def form():
    """Assessment form page"""
    return render_template('form.html')

@app.route('/result')
def result():
    """Result page"""
    return render_template('result.html')

# ==================== API Endpoints ====================

@app.route('/api/login', methods=['POST'])
def api_login():
    """Login API endpoint"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password required'}), 400
        
        if verify_user(username, password):
            return jsonify({
                'success': True,
                'message': 'Login successful',
                'username': username
            }), 200
        else:
            # Check if user exists
            if user_exists(username):
                return jsonify({
                    'success': False,
                    'message': 'Invalid password',
                    'action': 'retry'
                }), 401
            else:
                return jsonify({
                    'success': False,
                    'message': 'User not found',
                    'action': 'signup'
                }), 404
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/signup', methods=['POST'])
def api_signup():
    """Signup API endpoint"""
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        
        if not username or not password:
            return jsonify({'success': False, 'message': 'Username and password required'}), 400
        
        if len(username) < 3:
            return jsonify({'success': False, 'message': 'Username must be at least 3 characters'}), 400
        
        if len(password) < 6:
            return jsonify({'success': False, 'message': 'Password must be at least 6 characters'}), 400
        
        if user_exists(username):
            return jsonify({'success': False, 'message': 'Username already exists'}), 409
        
        if add_user(username, password):
            return jsonify({
                'success': True,
                'message': 'Account created successfully',
                'username': username
            }), 201
        else:
            return jsonify({'success': False, 'message': 'Failed to create account'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/predict', methods=['POST'])
def predict():
    """Main prediction endpoint"""
    try:
        if model is None:
            return jsonify({
                'success': False,
                'message': 'Model not loaded'
            }), 500
        
        data = request.get_json()
        
        # Extract features
        features = {}
        for col in feature_columns:
            if col not in data:
                return jsonify({
                    'success': False,
                    'message': f'Missing feature: {col}'
                }), 400
            try:
                features[col] = float(data[col])
            except (ValueError, TypeError):
                return jsonify({
                    'success': False,
                    'message': f'Invalid value for {col}'
                }), 400
        
        age = features.get('Age', 0.0)
        height = features.get('Height', 0.0)
        weight = features.get('Weight', 0.0)

        # ANN prediction first (for all ages).
        feature_vector = np.array([[features[col] for col in feature_columns]])
        feature_vector_scaled = scaler.transform(feature_vector)
        prediction = model.predict(feature_vector_scaled, verbose=0)
        predicted_class_idx = int(np.argmax(prediction[0]))
        confidence = float(np.max(prediction[0])) * 100

        predicted_raw = str(target_encoder.inverse_transform([predicted_class_idx])[0])
        canonical_label = NUMERIC_CLASS_FALLBACK.get(predicted_raw, predicted_raw)
        prediction_method = 'ann_model'
        note = ''

        all_predictions = {}
        for i, p in enumerate(prediction[0]):
            raw_label = str(target_encoder.classes_[i])
            canonical = NUMERIC_CLASS_FALLBACK.get(raw_label, raw_label)
            display = CLASS_DISPLAY_MAP.get(canonical, canonical.replace('_', ' '))
            all_predictions[display] = float(p) * 100

        # Pediatric fallback for ages 10-15 when ANN is uncertain or when BMI is in child low-BMI range.
        bmi = (weight / (height ** 2)) if height > 0 else 0
        if 10 <= age <= 15 and height > 0 and (confidence < 60 or bmi < 21.0):
            canonical_label, center_idx = _pediatric_class_from_bmi(bmi)
            probs = _pediatric_probabilities(center_idx)
            predicted_class_idx = int(np.argmax(probs))
            confidence = float(np.max(probs)) * 100
            prediction_method = 'pediatric_bmi_rule'
            note = 'Age 10-15 prediction used pediatric BMI logic for safer child-range classification.'

            all_predictions = {}
            ordered_labels = [
                'Insufficient_Weight',
                'Normal_Weight',
                'Overweight_Level_I',
                'Overweight_Level_II',
                'Obesity_Type_I',
                'Obesity_Type_II',
                'Obesity_Type_III'
            ]
            for i, canonical in enumerate(ordered_labels):
                display = CLASS_DISPLAY_MAP.get(canonical, canonical.replace('_', ' '))
                all_predictions[display] = float(probs[i]) * 100

        predicted_label = CLASS_DISPLAY_MAP.get(canonical_label, canonical_label.replace('_', ' '))
        risk_info = OBESITY_RISK_MAP.get(canonical_label, {'risk_level': 'Unknown', 'icon': '❓', 'color': '#95a5a6'})
        
        return jsonify({
            'success': True,
            'prediction': predicted_label,
            'prediction_index': int(predicted_class_idx),
            'confidence': round(confidence, 2),
            'risk_level': risk_info['risk_level'],
            'icon': risk_info['icon'],
            'color': risk_info['color'],
            'all_predictions': all_predictions,
            'prediction_method': prediction_method,
            'note': note
        }), 200
        
    except Exception as e:
        print(f"Prediction error: {e}")
        return jsonify({
            'success': False,
            'message': f'Prediction error: {str(e)}'
        }), 500

# ==================== Static Files ====================

@app.route('/css/<path:filename>')
def serve_css(filename):
    return send_from_directory(CSS_DIR, filename)

@app.route('/js/<path:filename>')
def serve_js(filename):
    return send_from_directory(JS_DIR, filename)

@app.route('/images/<path:filename>')
def serve_images(filename):
    return send_from_directory(IMAGES_DIR, filename)

# ==================== Error Handlers ====================

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Not found'}), 404

@app.errorhandler(500)
def server_error(error):
    return jsonify({'error': 'Server error'}), 500

# ==================== Main ====================

if __name__ == '__main__':
    print("Starting ObesiCare Flask API...")
    print("Available endpoints:")
    print("  GET  /               - Home page")
    print("  GET  /login          - Login page")
    print("  GET  /signup         - Sign up page")
    print("  GET  /form           - Assessment form")
    print("  GET  /result         - Result page")
    print("  POST /api/login      - Login API")
    print("  POST /api/signup     - Sign up API")
    print("  POST /predict        - Prediction API")
    
    # For development
    app.run(debug=True, host='0.0.0.0', port=5000)
